"""
===============================================================================
 Meridian Precast — Enhanced Analysis Workbook Generator
 ========================================================
 Creates a professional Excel workbook with:
   - Original data tabs (Labour Records, Jobs, Products, Work Centres)
   - Analysis tabs for each question (Q1-Q9)
   - Summary dashboard tab
   - Conditional formatting to highlight violations
   - Pivot-style tables for Question 8
===============================================================================
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side, 
                              numbers, NamedStyle)
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
import os

DATA_FILE   = r'Meridian Precast - Production Data.xlsx'
OUTPUT_FILE = r'Meridian Precast - Analysis Workbook.xlsx'

# ─── Load Data ───
labour   = pd.read_excel(DATA_FILE, sheet_name='Labour Records')
jobs     = pd.read_excel(DATA_FILE, sheet_name='Jobs')
products = pd.read_excel(DATA_FILE, sheet_name='Products')
centres  = pd.read_excel(DATA_FILE, sheet_name='Work Centres')

labour['Date'] = pd.to_datetime(labour['Date'])
jobs['Date Closed'] = pd.to_datetime(jobs['Date Closed'])

# ─── Styles ───
header_font    = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
header_fill    = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
title_font     = Font(name='Calibri', bold=True, size=14, color='2F5496')
subtitle_font  = Font(name='Calibri', bold=True, size=11, color='2F5496')
red_fill       = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
red_font       = Font(color='9C0006')
green_fill     = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
green_font     = Font(color='006100')
yellow_fill    = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
yellow_font    = Font(color='9C6500')
border         = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))

def style_header(ws, row=1, max_col=None):
    """Apply header styling to the first row."""
    if max_col is None:
        max_col = ws.max_column
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

def write_df(ws, df, start_row=1):
    """Write a DataFrame to a worksheet with formatting."""
    # Headers
    for c_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=c_idx, value=col_name)
    style_header(ws, row=start_row, max_col=len(df.columns))
    
    # Data
    for r_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            if pd.isna(val):
                cell.value = None
            elif isinstance(val, pd.Timestamp):
                cell.value = val.to_pydatetime()
                cell.number_format = 'YYYY-MM-DD'
            else:
                cell.value = val
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
    
    return start_row + len(df) + 1  # return next available row

def add_title(ws, title, row=1, col=1):
    """Add a title to a worksheet."""
    cell = ws.cell(row=row, column=col, value=title)
    cell.font = title_font
    return row + 2

# ─── Create Workbook ───
wb = Workbook()

# =============================================================================
# SUMMARY DASHBOARD TAB
# =============================================================================
ws_summary = wb.active
ws_summary.title = 'Analysis Summary'
ws_summary.sheet_properties.tabColor = '2F5496'

r = add_title(ws_summary, 'Meridian Precast — Data Quality Analysis Summary')

# Summary table
summary_data = [
    ['Q#', 'Rule', 'Question', 'Finding', 'Records/Items Affected'],
    ['Q1', 'R3', 'Missing operator records', 'YES — 2 records have no operator', 'Record 148, 175'],
    ['Q2', 'R6', 'Duplicate records', 'YES — 2 records are identical', 'Record 119, 120'],
    ['Q3', 'R1', 'Overtime (>8 hrs/day)', 'YES — 2 instances', 'OP-102 on 7/6, OP-104 on 6/8'],
    ['Q4', 'R4', 'Missing scrap reason', 'YES — 2 records, 14 units', 'Record 243, 250'],
    ['Q5', 'R2', 'Labour after job close', 'YES — 2 records on JOB-1014', 'Record 263, 266'],
    ['Q6', 'R5', 'Overproduction at Finishing', 'YES — JOB-1011 overproduced by 6', 'JOB-1011'],
    ['Q7', 'R8', 'Product priced below cost', 'YES — PRD-440 loses $54/unit', 'PRD-440'],
    ['Q8', 'R7', 'Labour cost by work centre', 'WC-350 at $0/hr is concerning', 'See Q8 tab'],
    ['Q9', '—', 'JOB-1006 vs JOB-1010', '59.32 hour difference', 'See Q9 tab'],
]

for row_data in summary_data:
    for c_idx, val in enumerate(row_data, 1):
        ws_summary.cell(row=r, column=c_idx, value=val).border = border
    r += 1

style_header(ws_summary, row=3, max_col=5)

# Apply conditional formatting to findings column
for row_idx in range(4, 13):
    cell = ws_summary.cell(row=row_idx, column=4)
    if cell.value and 'YES' in str(cell.value):
        cell.fill = red_fill
        cell.font = red_font

# Column widths
ws_summary.column_dimensions['A'].width = 6
ws_summary.column_dimensions['B'].width = 8
ws_summary.column_dimensions['C'].width = 32
ws_summary.column_dimensions['D'].width = 40
ws_summary.column_dimensions['E'].width = 30

# =============================================================================
# Q1 — MISSING OPERATOR
# =============================================================================
ws_q1 = wb.create_sheet('Q1 - Missing Operator')
ws_q1.sheet_properties.tabColor = 'FF0000'

r = add_title(ws_q1, 'Q1: Labour Records with Missing Operator (Rule R3)')
ws_q1.cell(row=r, column=1, value='Rule R3: Every labour record must name the operator who did the work.').font = subtitle_font
r += 2

missing_op = labour[labour['Operator ID'].isna() | labour['Operator Name'].isna()]
write_df(ws_q1, missing_op, start_row=r)

r += len(missing_op) + 3
ws_q1.cell(row=r, column=1, value=f'FINDING: {len(missing_op)} record(s) are missing operator information (Record IDs: {", ".join(map(str, missing_op["Record ID"]))}).')
ws_q1.cell(row=r, column=1).font = Font(bold=True, color='FF0000')

# =============================================================================
# Q2 — DUPLICATES
# =============================================================================
ws_q2 = wb.create_sheet('Q2 - Duplicates')
ws_q2.sheet_properties.tabColor = 'FF0000'

r = add_title(ws_q2, 'Q2: Duplicate Records (Rule R6)')
ws_q2.cell(row=r, column=1, value='Rule R6: Same operator should not have two identical records for the same job, stage and date.').font = subtitle_font
r += 2

dup_cols = ['Date', 'Job Number', 'Stage', 'Operator ID', 'Hours Worked', 'Units Completed', 'Units Scrapped']
dups = labour[labour.duplicated(subset=dup_cols, keep=False)].sort_values('Record ID')
write_df(ws_q2, dups, start_row=r)

r += len(dups) + 3
ws_q2.cell(row=r, column=1, value=f'FINDING: Records {", ".join(map(str, dups["Record ID"]))} are duplicates.').font = Font(bold=True, color='FF0000')
r += 1
ws_q2.cell(row=r, column=1, value='DETECTION: Group by (Date, Job, Stage, Operator, Hours, Units Completed, Units Scrapped) → flag groups with count > 1.').font = subtitle_font

# =============================================================================
# Q3 — OVERTIME
# =============================================================================
ws_q3 = wb.create_sheet('Q3 - Overtime')
ws_q3.sheet_properties.tabColor = 'FF9900'

r = add_title(ws_q3, 'Q3: Operators Exceeding 8 Hours/Day (Rule R1)')
ws_q3.cell(row=r, column=1, value='Rule R1: A shift is 8 hours. No operator should record more than 8 hours of work on one date.').font = subtitle_font
r += 2

daily = labour.groupby(['Operator ID', 'Operator Name', 'Date'])['Hours Worked'].sum().reset_index()
daily.columns = ['Operator ID', 'Operator Name', 'Date', 'Total Hours']
overtime = daily[daily['Total Hours'] > 8.0].sort_values('Total Hours', ascending=False)

write_df(ws_q3, overtime, start_row=r)

# Highlight overtime rows
for row_idx in range(r + 1, r + 1 + len(overtime)):
    for col_idx in range(1, 5):
        ws_q3.cell(row=row_idx, column=col_idx).fill = red_fill

r += len(overtime) + 3
for _, row_data in overtime.iterrows():
    op_name = row_data['Operator Name'] if pd.notna(row_data['Operator Name']) else 'MISSING'
    ws_q3.cell(row=r, column=1, 
               value=f'FINDING: {row_data["Operator ID"]} ({op_name}) worked {row_data["Total Hours"]:.2f} hours on {row_data["Date"].strftime("%Y-%m-%d")}.')
    ws_q3.cell(row=r, column=1).font = Font(bold=True, color='FF0000')
    r += 1

# =============================================================================
# Q4 — MISSING SCRAP REASON
# =============================================================================
ws_q4 = wb.create_sheet('Q4 - Missing Scrap Reason')
ws_q4.sheet_properties.tabColor = 'FF0000'

r = add_title(ws_q4, 'Q4: Scrapped Units Without Scrap Reason (Rule R4)')
ws_q4.cell(row=r, column=1, value='Rule R4: If any units are scrapped, a scrap reason must be recorded.').font = subtitle_font
r += 2

scrap_no_reason = labour[(labour['Units Scrapped'] > 0) & (labour['Scrap Reason'].isna())]
write_df(ws_q4, scrap_no_reason, start_row=r)

r += len(scrap_no_reason) + 3
total_units = int(scrap_no_reason['Units Scrapped'].sum())
ws_q4.cell(row=r, column=1, 
           value=f'FINDING: {len(scrap_no_reason)} record(s) have scrapped units but no scrap reason. Total units affected: {total_units}.')
ws_q4.cell(row=r, column=1).font = Font(bold=True, color='FF0000')

# =============================================================================
# Q5 — POST-CLOSE LABOUR
# =============================================================================
ws_q5 = wb.create_sheet('Q5 - Post-Close Labour')
ws_q5.sheet_properties.tabColor = 'FF9900'

r = add_title(ws_q5, 'Q5: Labour Records Dated After Job Closure (Rule R2)')
ws_q5.cell(row=r, column=1, value='Rule R2: No labour should be dated after a job\'s Date Closed.').font = subtitle_font
r += 2

labour_jobs = labour.merge(jobs[['Job Number', 'Date Closed']], on='Job Number', how='left')
post_close = labour_jobs[
    (labour_jobs['Date Closed'].notna()) & (labour_jobs['Date'] > labour_jobs['Date Closed'])
]
post_close_display = post_close[['Record ID', 'Date', 'Job Number', 'Stage', 'Operator Name', 
                                  'Hours Worked', 'Date Closed']].copy()
post_close_display['Days After Close'] = (post_close_display['Date'] - post_close_display['Date Closed']).dt.days
write_df(ws_q5, post_close_display, start_row=r)

r += len(post_close_display) + 3
ws_q5.cell(row=r, column=1, 
           value=f'FINDING: {len(post_close)} record(s) have labour dated after the job was closed. Job: {", ".join(post_close["Job Number"].unique())}.')
ws_q5.cell(row=r, column=1).font = Font(bold=True, color='FF0000')

# =============================================================================
# Q6 — OVERPRODUCTION
# =============================================================================
ws_q6 = wb.create_sheet('Q6 - Overproduction')
ws_q6.sheet_properties.tabColor = 'FF9900'

r = add_title(ws_q6, 'Q6: Jobs Finishing More Units Than Ordered (Rule R5)')
ws_q6.cell(row=r, column=1, value='Rule R5: Units completed at the Finishing stage should not be more than the quantity ordered.').font = subtitle_font
r += 2

finishing = labour[labour['Stage'] == 'Finishing'].groupby('Job Number')['Units Completed'].sum().reset_index()
finishing.columns = ['Job Number', 'Total Finished']
finishing_vs = finishing.merge(jobs[['Job Number', 'Quantity Ordered', 'Product Code']], on='Job Number')
finishing_vs['Overage'] = finishing_vs['Total Finished'] - finishing_vs['Quantity Ordered']
finishing_vs['Status'] = finishing_vs['Overage'].apply(lambda x: '⚠ OVERPRODUCED' if x > 0 else ('OK' if x == 0 else 'Under'))
finishing_vs = finishing_vs[['Job Number', 'Product Code', 'Quantity Ordered', 'Total Finished', 'Overage', 'Status']]

next_r = write_df(ws_q6, finishing_vs.sort_values('Job Number'), start_row=r)

# Highlight overproduced rows
for row_idx in range(r + 1, r + 1 + len(finishing_vs)):
    status_cell = ws_q6.cell(row=row_idx, column=6)
    if status_cell.value and 'OVER' in str(status_cell.value):
        for col_idx in range(1, 7):
            ws_q6.cell(row=row_idx, column=col_idx).fill = red_fill

r = next_r + 1
overproduced = finishing_vs[finishing_vs['Overage'] > 0]
for _, row_data in overproduced.iterrows():
    ws_q6.cell(row=r, column=1, 
               value=f'FINDING: {row_data["Job Number"]} finished {int(row_data["Total Finished"])} units but only {int(row_data["Quantity Ordered"])} were ordered (overage: {int(row_data["Overage"])}).')
    ws_q6.cell(row=r, column=1).font = Font(bold=True, color='FF0000')
    r += 1

# =============================================================================
# Q7 — PRODUCT ISSUES
# =============================================================================
ws_q7 = wb.create_sheet('Q7 - Product Issues')
ws_q7.sheet_properties.tabColor = 'FF0000'

r = add_title(ws_q7, 'Q7: Product Margin Analysis (Rule R8)')
ws_q7.cell(row=r, column=1, value='Rule R8: Every product should sell for more than it costs to make.').font = subtitle_font
r += 2

prod_analysis = products.copy()
prod_analysis['Margin ($)'] = prod_analysis['Selling Price per Unit ($)'] - prod_analysis['Cost to Make per Unit ($)']
prod_analysis['Margin (%)'] = (prod_analysis['Margin ($)'] / prod_analysis['Selling Price per Unit ($)'] * 100).round(1)
prod_analysis['Status'] = prod_analysis['Margin ($)'].apply(lambda x: '❌ LOSS' if x < 0 else '✅ Profitable')

next_r = write_df(ws_q7, prod_analysis, start_row=r)

# Highlight loss-making product
for row_idx in range(r + 1, r + 1 + len(prod_analysis)):
    status_cell = ws_q7.cell(row=row_idx, column=7)
    if status_cell.value and 'LOSS' in str(status_cell.value):
        for col_idx in range(1, 8):
            ws_q7.cell(row=row_idx, column=col_idx).fill = red_fill

r = next_r + 1
loss_prods = prod_analysis[prod_analysis['Margin ($)'] < 0]
for _, row_data in loss_prods.iterrows():
    ws_q7.cell(row=r, column=1, 
               value=f'FINDING: {row_data["Product Code"]} ({row_data["Description"]}) costs ${row_data["Cost to Make per Unit ($)"]:.2f} to make but sells for ${row_data["Selling Price per Unit ($)"]:.2f}. The company loses ${abs(row_data["Margin ($)"]):.2f} on every unit.')
    ws_q7.cell(row=r, column=1).font = Font(bold=True, color='FF0000')
    r += 1

# =============================================================================
# Q8 — LABOUR COST BY WORK CENTRE
# =============================================================================
ws_q8 = wb.create_sheet('Q8 - Labour Cost')
ws_q8.sheet_properties.tabColor = '2F5496'

r = add_title(ws_q8, 'Q8: Total Hours and Labour Cost by Work Centre (Rule R7)')
ws_q8.cell(row=r, column=1, value='Rule R7: Labour cost for a record = Hours Worked × the Labour Rate for that work centre.').font = subtitle_font
r += 2

labour_wc = labour.merge(centres, on='Work Centre', how='left')
labour_wc['Labour Cost ($)'] = labour_wc['Hours Worked'] * labour_wc['Labour Rate ($ per hour)']

wc_summary = labour_wc.groupby(['Work Centre', 'Name']).agg(
    Records=('Record ID', 'count'),
    Total_Hours=('Hours Worked', 'sum'),
    Labour_Rate=('Labour Rate ($ per hour)', 'first'),
    Total_Labour_Cost=('Labour Cost ($)', 'sum')
).reset_index()

wc_summary['Total_Hours'] = wc_summary['Total_Hours'].round(2)
wc_summary['Total_Labour_Cost'] = wc_summary['Total_Labour_Cost'].round(2)
wc_summary.columns = ['Work Centre', 'Name', 'Records', 'Total Hours', 'Rate ($/hr)', 'Total Labour Cost ($)']

next_r = write_df(ws_q8, wc_summary, start_row=r)

# Add totals row
total_row = next_r - 1
ws_q8.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
ws_q8.cell(row=total_row, column=3, value=int(wc_summary['Records'].sum())).font = Font(bold=True)
ws_q8.cell(row=total_row, column=4, value=round(wc_summary['Total Hours'].sum(), 2)).font = Font(bold=True)
ws_q8.cell(row=total_row, column=6, value=round(wc_summary['Total Labour Cost ($)'].sum(), 2)).font = Font(bold=True)
for col in range(1, 7):
    ws_q8.cell(row=total_row, column=col).border = Border(top=Side(style='double'), bottom=Side(style='double'))

# Highlight the $0 rate row
for row_idx in range(r + 1, total_row):
    rate_cell = ws_q8.cell(row=row_idx, column=5)
    if rate_cell.value == 0:
        for col_idx in range(1, 7):
            ws_q8.cell(row=row_idx, column=col_idx).fill = yellow_fill

r = total_row + 2
ws_q8.cell(row=r, column=1, value='CONCERN: WC-350 (Finishing) has a labour rate of $0.00/hr.').font = Font(bold=True, color='FF9900')
r += 1
ws_q8.cell(row=r, column=1, value='214.56 hours of Finishing work contribute $0.00 to labour costs.').font = subtitle_font
r += 1
ws_q8.cell(row=r, column=1, value='This may mean: (1) Salaried staff, (2) Automated process, or (3) Data entry error.').font = subtitle_font

# Add a bar chart
chart = BarChart()
chart.type = 'col'
chart.title = 'Total Labour Cost by Work Centre'
chart.y_axis.title = 'Labour Cost ($)'
chart.x_axis.title = 'Work Centre'
data_ref = Reference(ws_q8, min_col=6, min_row=r-len(wc_summary)-4, max_row=r-5)
cats_ref = Reference(ws_q8, min_col=2, min_row=r-len(wc_summary)-3, max_row=r-5)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.width = 18
chart.height = 12
ws_q8.add_chart(chart, f'A{r+2}')

# =============================================================================
# Q9 — JOB COMPARISON
# =============================================================================
ws_q9 = wb.create_sheet('Q9 - Job Comparison')
ws_q9.sheet_properties.tabColor = '2F5496'

r = add_title(ws_q9, 'Q9: JOB-1006 vs JOB-1010 — Labour Hours Comparison')
r += 1

job_1006 = labour[labour['Job Number'] == 'JOB-1006']
job_1010 = labour[labour['Job Number'] == 'JOB-1010']

hrs_1006 = job_1006['Hours Worked'].sum()
hrs_1010 = job_1010['Hours Worked'].sum()
diff = abs(hrs_1006 - hrs_1010)

# Comparison table
comp_data = [
    ['Metric', 'JOB-1006', 'JOB-1010', 'Difference'],
    ['Product', 'PRD-110', 'PRD-110', '—'],
    ['Quantity Ordered', 24, 24, '—'],
    ['Total Labour Hours', round(hrs_1006, 2), round(hrs_1010, 2), round(diff, 2)],
    ['Number of Records', len(job_1006), len(job_1010), abs(len(job_1006) - len(job_1010))],
    ['Total Units Completed', int(job_1006['Units Completed'].sum()), int(job_1010['Units Completed'].sum()), ''],
    ['Total Units Scrapped', int(job_1006['Units Scrapped'].sum()), int(job_1010['Units Scrapped'].sum()), ''],
    ['Date Range', f"{job_1006['Date'].min().strftime('%m/%d')} - {job_1006['Date'].max().strftime('%m/%d')}", 
                   f"{job_1010['Date'].min().strftime('%m/%d')} - {job_1010['Date'].max().strftime('%m/%d')}", ''],
    ['Calendar Days', (job_1006['Date'].max()-job_1006['Date'].min()).days+1, 
                      (job_1010['Date'].max()-job_1010['Date'].min()).days+1, ''],
]

for row_data in comp_data:
    for c_idx, val in enumerate(row_data, 1):
        ws_q9.cell(row=r, column=c_idx, value=val).border = border
    r += 1

style_header(ws_q9, row=4, max_col=4)
ws_q9.column_dimensions['A'].width = 25
ws_q9.column_dimensions['B'].width = 18
ws_q9.column_dimensions['C'].width = 18
ws_q9.column_dimensions['D'].width = 15

# Highlight the hours difference row
for col in range(1, 5):
    ws_q9.cell(row=7, column=col).fill = yellow_fill

# Stage breakdown
r += 2
ws_q9.cell(row=r, column=1, value='Hours Breakdown by Stage').font = subtitle_font
r += 1
stage_header = ['Stage', 'JOB-1006 Hours', 'JOB-1010 Hours', 'Difference']
for c_idx, val in enumerate(stage_header, 1):
    ws_q9.cell(row=r, column=c_idx, value=val).border = border
style_header(ws_q9, row=r, max_col=4)
r += 1

for stage in ['Mixing', 'Casting', 'Curing', 'Finishing']:
    h1006 = job_1006[job_1006['Stage'] == stage]['Hours Worked'].sum()
    h1010 = job_1010[job_1010['Stage'] == stage]['Hours Worked'].sum()
    row_vals = [stage, round(h1006, 2), round(h1010, 2), round(abs(h1010 - h1006), 2)]
    for c_idx, val in enumerate(row_vals, 1):
        ws_q9.cell(row=r, column=c_idx, value=val).border = border
    # Highlight Casting row (biggest difference)
    if stage == 'Casting':
        for col in range(1, 5):
            ws_q9.cell(row=r, column=col).fill = red_fill
    r += 1

# Analysis
r += 2
ws_q9.cell(row=r, column=1, value='ANALYSIS').font = title_font
r += 1
ws_q9.cell(row=r, column=1, value=f'The difference is {diff:.2f} hours — JOB-1010 used nearly DOUBLE the labour of JOB-1006.').font = Font(bold=True)
r += 2

hypotheses = [
    '1. OPERATOR EXPERIENCE: JOB-1010 includes OP-106 (Singh, A.) who does not appear on any other job.',
    '   This may be a new or less experienced operator. OP-106 logged 56.00 hours (all exactly 8.00/day),',
    '   accounting for most of the difference. The Casting stage alone shows a 58.61-hour gap.',
    '',
    '2. HIGHER SCRAP / REWORK: JOB-1010 scrapped 29 units vs. only 4 for JOB-1006.',
    '   More scrap means more rework and extra hours to produce replacement units.',
    '',
    '3. LONGER DURATION: JOB-1010 spanned 24 calendar days vs. 18 for JOB-1006,',
    '   possibly due to production interruptions, machine downtime, or scheduling issues.',
    '',
    'NEXT STEPS:',
    '  - Ask the Production Manager whether OP-106 was being trained during JOB-1010.',
    '  - Check if there were equipment issues or downtime during JOB-1010.',
    '  - Investigate whether the 8.00-hour entries from OP-106 are actual or estimated.',
]

for line in hypotheses:
    ws_q9.cell(row=r, column=1, value=line)
    r += 1

# =============================================================================
# ORIGINAL DATA TABS (for reference)
# =============================================================================
ws_labour = wb.create_sheet('Labour Records')
write_df(ws_labour, labour)

ws_jobs = wb.create_sheet('Jobs')
write_df(ws_jobs, jobs)

ws_products = wb.create_sheet('Products')
write_df(ws_products, products)

ws_centres = wb.create_sheet('Work Centres')
write_df(ws_centres, centres)

# ─── Auto-fit column widths (approximate) ───
for ws in wb.worksheets:
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

# ─── Save ───
wb.save(OUTPUT_FILE)
print(f"✓ Analysis workbook saved as: {OUTPUT_FILE}")
print(f"  Tabs: {', '.join([ws.title for ws in wb.worksheets])}")
